#!/usr/bin/env python3
"""
AK System Cleaner v4.3
Windows System Maintenance & Optimization Tool
Developed BY Ahmad_Khaled
"""

import tkinter as tk
from tkinter import messagebox, filedialog
import subprocess
import threading
import os
import sys
import ctypes
import time
import datetime
import traceback

# ============================================================
# Check Admin
# ============================================================
def is_admin():
    try:
        return ctypes.windll.shell32.IsUserAnAdmin()
    except:
        return False

if not is_admin():
    try:
        ctypes.windll.shell32.ShellExecuteW(
            None, "runas", sys.executable, " ".join([f'"{a}"' for a in sys.argv]), None, 1
        )
        sys.exit(0)
    except:
        pass

# ============================================================
# Step Definitions
# ============================================================
STEPS = [
    {
        "id": "virus",
        "icon": "\U0001F6E1",
        "name": "Remove Virus Remnants",
        "desc": "بيمسح فيروسات USB القديمة من كل الدرايفات",
        "default": True,
    },
    {
        "id": "temp",
        "icon": "\U0001F9F9",
        "name": "Clean Temp Files",
        "desc": "بيحذف الملفات المؤقتة وبيحرر مساحة",
        "default": True,
    },
    {
        "id": "recycle",
        "icon": "\U0001F5D1",
        "name": "Empty Recycle Bin",
        "desc": "بيفضي سلة المحذوفات تلقائي",
        "default": True,
    },
    {
        "id": "dns",
        "icon": "\U0001F310",
        "name": "Clear DNS Cache",
        "desc": "بيسرع التصفح عن طريق مسح DNS القديم",
        "default": True,
    },
    {
        "id": "hidden",
        "icon": "\U0001F441",
        "name": "Show Hidden Files",
        "desc": "بيظهر الملفات المخفية في Explorer",
        "default": True,
    },
    {
        "id": "darkmode",
        "icon": "\U0001F319",
        "name": "Enable Dark Mode",
        "desc": "بيحول الويندوز كله لـ Dark Mode",
        "default": False,
    },
    {
        "id": "telemetry",
        "icon": "\U0001F512",
        "name": "Disable Telemetry",
        "desc": "بيوقف ارسال بيانات الاستخدام لـ Microsoft",
        "default": False,
    },
    {
        "id": "startup",
        "icon": "\U0001F680",
        "name": "Disable Startup Programs",
        "desc": "بيوقف البرامج اللي بتشتغل مع بدء التشغيل",
        "default": False,
    },
    {
        "id": "chkdsk",
        "icon": "\U0001F4BF",
        "name": "Schedule CHKDSK",
        "desc": "بيجدول فحص الهارد لاصلاح الاخطاء مع اول ريستارت",
        "default": True,
    },
    {
        "id": "sfc",
        "icon": "\u2699",
        "name": "SFC Scan",
        "desc": "بيفحص ملفات النظام ويصلح التالف منها (5-15 دقيقة)",
        "default": True,
    },
    {
        "id": "dism",
        "icon": "\U0001F3D7",
        "name": "DISM Repair",
        "desc": "بيصلح صورة الويندوز لو SFC مش قدر يصلحها",
        "default": True,
    },
    {
        "id": "explorer",
        "icon": "\U0001F504",
        "name": "Restart Explorer",
        "desc": "بيعيد تشغيل Explorer عشان التغييرات تظهر فورا",
        "default": True,
    },
]

USB_STEP = {
    "id": "usb",
    "icon": "\U0001F4F1",
    "name": "USB Cleaner",
    "desc": "بيكتشف USB المتوصلة وبيمسح فيروساتها (بدون فورمات)",
    "default": False,
}

# ============================================================
# Colors
# ============================================================
BG = "#0D0D0D"
BG2 = "#161616"
CARD = "#1A1A1A"
ACCENT = "#00C896"
TEXT = "#E0E0E0"
DIM = "#666666"
DESC_CLR = "#707070"
RED = "#FF4B4B"
YELLOW = "#FFD93D"
BLUE = "#4BA3FF"

# ============================================================
# Main Application
# ============================================================
class AKCleanerApp:
    def __init__(self):
        self.root = tk.Tk()
        self.root.title("AK System Cleaner v4.3 — BY Ahmad_Khaled")
        self.root.configure(bg=BG)
        self.root.resizable(True, True)

        # Set window size
        w, h = 920, 900
        sw = self.root.winfo_screenwidth()
        sh = self.root.winfo_screenheight()
        x = (sw - w) // 2
        y = (sh - h) // 2
        self.root.geometry(f"{w}x{h}+{x}+{y}")

        self.checks = {}
        self.running = False
        self.log_lines = []
        self.start_time = None

        self._build_ui()

    def _build_ui(self):
        # ============== HEADER ==============
        header = tk.Frame(self.root, bg=BG)
        header.pack(fill="x", padx=16, pady=(12, 2))

        tk.Label(
            header,
            text="\U0001F6E0  AK System Cleaner",
            font=("Segoe UI", 16, "bold"),
            fg=ACCENT,
            bg=BG,
        ).pack(side="left")

        tk.Label(
            header,
            text="v4.3",
            font=("Segoe UI", 10),
            fg=DIM,
            bg=BG,
        ).pack(side="left", padx=(6, 0), pady=(4, 0))

        # User & PC info
        try:
            user = os.environ.get("USERNAME", "User")
            pc = os.environ.get("COMPUTERNAME", "PC")
            info_text = f"{user}@{pc}"
        except:
            info_text = ""

        tk.Label(
            header,
            text=info_text,
            font=("Consolas", 9),
            fg=DIM,
            bg=BG,
        ).pack(side="right")

        # ============== DEVELOPER CREDIT ==============
        dev_frame = tk.Frame(self.root, bg=BG)
        dev_frame.pack(fill="x", padx=16, pady=(0, 2))

        tk.Label(
            dev_frame,
            text="Developed BY Ahmad_Khaled",
            font=("Segoe UI", 9, "italic"),
            fg=ACCENT,
            bg=BG,
        ).pack(side="left")

        # ============== SEPARATOR ==============
        tk.Frame(self.root, bg=ACCENT, height=1).pack(fill="x", padx=16, pady=(2, 8))

        # ============== MAIN CONTENT ==============
        main = tk.Frame(self.root, bg=BG)
        main.pack(fill="both", expand=True, padx=16, pady=0)

        # ---------- Left: Steps ----------
        left = tk.Frame(main, bg=BG)
        left.pack(side="left", fill="both", expand=True, padx=(0, 8))

        tk.Label(
            left,
            text="Steps",
            font=("Segoe UI", 11, "bold"),
            fg=TEXT,
            bg=BG,
        ).pack(anchor="w", pady=(0, 4))

        steps_frame = tk.Frame(left, bg=BG)
        steps_frame.pack(fill="both", expand=True)

        # Canvas for scrolling
        canvas = tk.Canvas(steps_frame, bg=BG, highlightthickness=0)
        scrollbar = tk.Scrollbar(steps_frame, orient="vertical", command=canvas.yview)
        scroll_frame = tk.Frame(canvas, bg=BG)

        scroll_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )

        canvas.create_window((0, 0), window=scroll_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)

        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        # Mouse wheel
        def _on_mousewheel(event):
            canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
        canvas.bind_all("<MouseWheel>", _on_mousewheel)

        # Add main steps
        for step in STEPS:
            self._add_step_card(scroll_frame, step)

        # USB section
        tk.Frame(scroll_frame, bg=ACCENT, height=1).pack(fill="x", pady=(8, 4))
        tk.Label(
            scroll_frame,
            text="USB Tools",
            font=("Segoe UI", 9, "bold"),
            fg=ACCENT,
            bg=BG,
        ).pack(anchor="w")
        self._add_step_card(scroll_frame, USB_STEP)

        # ---------- Right: Output ----------
        right = tk.Frame(main, bg=BG)
        right.pack(side="right", fill="both", expand=True, padx=(8, 0))

        tk.Label(
            right,
            text="Live Output",
            font=("Segoe UI", 11, "bold"),
            fg=TEXT,
            bg=BG,
        ).pack(anchor="w", pady=(0, 4))

        self.log_box = tk.Text(
            right,
            bg="#111111",
            fg="#CCCCCC",
            font=("Consolas", 9),
            relief="flat",
            wrap="word",
            state="disabled",
            padx=8,
            pady=8,
        )
        self.log_box.pack(fill="both", expand=True)

        # Text color tags
        self.log_box.tag_configure("green", foreground="#00C896")
        self.log_box.tag_configure("yellow", foreground="#FFD93D")
        self.log_box.tag_configure("red", foreground="#FF4B4B")
        self.log_box.tag_configure("blue", foreground="#4BA3FF")
        self.log_box.tag_configure("dim", foreground="#666666")

        # ============== PROGRESS ==============
        prog_frame = tk.Frame(self.root, bg=BG)
        prog_frame.pack(fill="x", padx=16, pady=(8, 4))

        self.step_label = tk.Label(
            prog_frame,
            text="Ready",
            font=("Segoe UI", 9),
            fg=DIM,
            bg=BG,
        )
        self.step_label.pack(anchor="w")

        self.progress_canvas = tk.Canvas(
            prog_frame, bg="#1A1A1A", height=8, highlightthickness=0
        )
        self.progress_canvas.pack(fill="x", pady=(2, 0))

        self.pct_label = tk.Label(
            prog_frame,
            text="0%",
            font=("Consolas", 9),
            fg=DIM,
            bg=BG,
        )
        self.pct_label.pack(anchor="e")

        # ============== BUTTONS ==============
        btn_frame = tk.Frame(self.root, bg=BG)
        btn_frame.pack(fill="x", padx=16, pady=(4, 6))

        self.run_btn = tk.Button(
            btn_frame,
            text="\u25B6  Run Cleaner",
            font=("Segoe UI", 11, "bold"),
            fg="#000000",
            bg=ACCENT,
            activebackground="#00A87A",
            activeforeground="#000000",
            relief="flat",
            cursor="hand2",
            padx=20,
            pady=6,
            command=self._run,
        )
        self.run_btn.pack(side="left")

        tk.Button(
            btn_frame,
            text="\U0001F4CB  Export Log",
            font=("Segoe UI", 9),
            fg=TEXT,
            bg="#2A2A2A",
            activebackground="#3A3A3A",
            relief="flat",
            cursor="hand2",
            padx=12,
            pady=4,
            command=self._export_log,
        ).pack(side="left", padx=(8, 0))

        tk.Button(
            btn_frame,
            text="Clear",
            font=("Segoe UI", 9),
            fg=DIM,
            bg="#1A1A1A",
            activebackground="#2A2A2A",
            relief="flat",
            cursor="hand2",
            padx=12,
            pady=4,
            command=self._clear_log,
        ).pack(side="left", padx=(8, 0))

        # Timer
        self.timer_label = tk.Label(
            btn_frame,
            text="",
            font=("Consolas", 9),
            fg=DIM,
            bg=BG,
        )
        self.timer_label.pack(side="right")

        # ============== FOOTER CREDIT ==============
        footer = tk.Frame(self.root, bg=BG)
        footer.pack(fill="x", padx=16, pady=(2, 10))

        tk.Frame(footer, bg="#222222", height=1).pack(fill="x", pady=(0, 4))

        tk.Label(
            footer,
            text="\u00A9 2025 AK System Cleaner — Developed BY Ahmad_Khaled",
            font=("Segoe UI", 8),
            fg=DIM,
            bg=BG,
        ).pack(side="left")

        tk.Label(
            footer,
            text="All Rights Reserved",
            font=("Segoe UI", 8),
            fg="#444444",
            bg=BG,
        ).pack(side="right")

    # ============================================================
    # Step Card
    # ============================================================
    def _add_step_card(self, parent, step):
        card = tk.Frame(parent, bg=CARD, padx=8, pady=4)
        card.pack(fill="x", pady=2)

        var = tk.BooleanVar(value=step["default"])
        self.checks[step["id"]] = var

        top = tk.Frame(card, bg=CARD)
        top.pack(fill="x")

        cb = tk.Checkbutton(
            top,
            variable=var,
            bg=CARD,
            activebackground=CARD,
            selectcolor="#2A2A2A",
            relief="flat",
        )
        cb.pack(side="left")

        tk.Label(
            top,
            text=f'{step["icon"]}  {step["name"]}',
            font=("Segoe UI", 9, "bold"),
            fg=TEXT,
            bg=CARD,
        ).pack(side="left")

        tk.Label(
            card,
            text=step["desc"],
            font=("Segoe UI", 8),
            fg=DESC_CLR,
            bg=CARD,
        ).pack(anchor="w", padx=(24, 0))

    # ============================================================
    # Logging
    # ============================================================
    def _log(self, text, tag=""):
        self.log_lines.append((text, tag))
        def _update():
            self.log_box.configure(state="normal")
            if tag:
                self.log_box.insert("end", text + "\n", tag)
            else:
                self.log_box.insert("end", text + "\n")
            self.log_box.see("end")
            self.log_box.configure(state="disabled")
        self.root.after(0, _update)

    def _update_progress(self, current, total, step_name=""):
        pct = int((current / total) * 100) if total > 0 else 0
        self.pct_label.configure(text=f"{pct}%")
        self.step_label.configure(text=step_name, fg=YELLOW)

        self.progress_canvas.delete("all")
        w = self.progress_canvas.winfo_width()
        if w > 1:
            fill_w = int(w * pct / 100)
            self.progress_canvas.create_rectangle(
                0, 0, fill_w, 8, fill=ACCENT, outline=""
            )

    # ============================================================
    # Run Command Helper
    # ============================================================
    def _cmd(self, command, shell=True):
        try:
            result = subprocess.run(
                command,
                shell=shell,
                capture_output=True,
                text=True,
                timeout=300,
            )
            return result.returncode == 0, result.stdout.strip()
        except subprocess.TimeoutExpired:
            return False, "Timeout"
        except Exception as e:
            return False, str(e)

    # ============================================================
    # Step Implementations
    # ============================================================
    def _step_virus(self):
        self._log("=== Remove Virus Remnants ===", "blue")
        targets = [
            "autorun.inf", "copy.exe", "host.exe", "rose.exe",
            "svchost.exe", "xcopy.exe", "AdobeR.exe", "Thumbs.db",
        ]
        import string
        count = 0
        for d in string.ascii_uppercase:
            drive = f"{d}:\\"
            if os.path.exists(drive):
                for t in targets:
                    fp = os.path.join(drive, t)
                    if os.path.exists(fp):
                        try:
                            os.remove(fp)
                            self._log(f"  Removed: {fp}", "green")
                            count += 1
                        except:
                            self._log(f"  Cannot remove: {fp}", "red")
        if count == 0:
            self._log("  No virus remnants found", "dim")
        self._log(f"  Done — removed {count} files", "green")

    def _step_temp(self):
        self._log("=== Clean Temp Files ===", "blue")
        dirs = [
            os.environ.get("TEMP", ""),
            os.path.join(os.environ.get("WINDIR", "C:\\Windows"), "Temp"),
            os.path.join(os.environ.get("WINDIR", "C:\\Windows"), "Prefetch"),
        ]
        total = 0
        for d in dirs:
            if os.path.exists(d):
                for root_dir, dirs_list, files in os.walk(d):
                    for f in files:
                        try:
                            fp = os.path.join(root_dir, f)
                            os.remove(fp)
                            total += 1
                        except:
                            pass
        self._log(f"  Deleted {total} temp files", "green")

    def _step_recycle(self):
        self._log("=== Empty Recycle Bin ===", "blue")
        try:
            ctypes.windll.shell32.SHEmptyRecycleBinW(None, None, 7)
            self._log("  Recycle Bin emptied", "green")
        except:
            self._log("  Already empty or error", "dim")

    def _step_dns(self):
        self._log("=== Clear DNS Cache ===", "blue")
        ok, out = self._cmd("ipconfig /flushdns")
        if ok:
            self._log("  DNS cache cleared", "green")
        else:
            self._log(f"  Error: {out}", "red")

    def _step_hidden(self):
        self._log("=== Show Hidden Files ===", "blue")
        cmds = [
            'reg add "HKCU\\Software\\Microsoft\\Windows\\CurrentVersion\\Explorer\\Advanced" /v Hidden /t REG_DWORD /d 1 /f',
            'reg add "HKCU\\Software\\Microsoft\\Windows\\CurrentVersion\\Explorer\\Advanced" /v ShowSuperHidden /t REG_DWORD /d 1 /f',
        ]
        for c in cmds:
            self._cmd(c)
        self._log("  Hidden files now visible", "green")

    def _step_darkmode(self):
        self._log("=== Enable Dark Mode ===", "blue")
        cmds = [
            'reg add "HKCU\\Software\\Microsoft\\Windows\\CurrentVersion\\Themes\\Personalize" /v AppsUseLightTheme /t REG_DWORD /d 0 /f',
            'reg add "HKCU\\Software\\Microsoft\\Windows\\CurrentVersion\\Themes\\Personalize" /v SystemUsesLightTheme /t REG_DWORD /d 0 /f',
        ]
        for c in cmds:
            self._cmd(c)
        self._log("  Dark Mode enabled", "green")

    def _step_telemetry(self):
        self._log("=== Disable Telemetry ===", "blue")
        cmds = [
            'reg add "HKLM\\SOFTWARE\\Policies\\Microsoft\\Windows\\DataCollection" /v AllowTelemetry /t REG_DWORD /d 0 /f',
            'sc config DiagTrack start= disabled',
            'sc stop DiagTrack',
        ]
        for c in cmds:
            self._cmd(c)
        self._log("  Telemetry disabled", "green")

    def _step_startup(self):
        self._log("=== Disable Startup Programs ===", "blue")
        self._log("  Listing startup items...", "yellow")
        ok, out = self._cmd(
            'powershell -Command "Get-CimInstance Win32_StartupCommand | Select-Object Name, Command | Format-Table -AutoSize"'
        )
        if ok and out:
            for line in out.split("\n")[:15]:
                if line.strip():
                    self._log(f"  {line.strip()}", "dim")
        self._log("  Use Task Manager > Startup to disable specific apps", "yellow")

    def _step_chkdsk(self):
        self._log("=== Schedule CHKDSK ===", "blue")
        ok, out = self._cmd('echo Y | chkdsk C: /F /R /X')
        self._log("  CHKDSK scheduled for next reboot", "green")

    def _step_sfc(self):
        self._log("=== SFC Scan (may take 5-15 min) ===", "blue")
        self._log("  Running sfc /scannow ...", "yellow")
        ok, out = self._cmd("sfc /scannow")
        if ok:
            self._log("  SFC completed successfully", "green")
        else:
            self._log("  SFC finished with issues", "yellow")
        if out:
            for line in out.split("\n")[-5:]:
                if line.strip():
                    self._log(f"  {line.strip()}", "dim")

    def _step_dism(self):
        self._log("=== DISM Repair (may take a while) ===", "blue")
        self._log("  Running DISM...", "yellow")
        ok, out = self._cmd("DISM /Online /Cleanup-Image /RestoreHealth")
        if ok:
            self._log("  DISM completed successfully", "green")
        else:
            self._log("  DISM finished with issues", "yellow")

    def _step_explorer(self):
        self._log("=== Restart Explorer ===", "blue")
        self._cmd("taskkill /f /im explorer.exe")
        time.sleep(1)
        subprocess.Popen("explorer.exe", shell=True)
        self._log("  Explorer restarted", "green")

    def _step_usb(self):
        self._log("=== USB Cleaner (No Format) ===", "blue")
        import string as string_mod
        targets = [
            "autorun.inf", "copy.exe", "host.exe", "rose.exe",
            "svchost.exe", "Thumbs.db",
        ]
        found_usb = False
        for d in string_mod.ascii_uppercase:
            drive = f"{d}:\\"
            if not os.path.exists(drive):
                continue
            try:
                dtype = ctypes.windll.kernel32.GetDriveTypeW(drive)
                if dtype != 2:  # DRIVE_REMOVABLE
                    continue
            except:
                continue

            found_usb = True
            self._log(f"  Found USB: {drive}", "green")

            count = 0
            for t in targets:
                fp = os.path.join(drive, t)
                if os.path.exists(fp):
                    try:
                        os.remove(fp)
                        count += 1
                        self._log(f"    Removed: {t}", "green")
                    except:
                        self._log(f"    Cannot remove: {t}", "red")

            # Reset hidden attributes
            self._cmd(f'attrib -h -s -r "{drive}*.*" /s /d')
            self._log(f"    Cleaned {count} virus files, attributes reset", "green")

        if not found_usb:
            self._log("  No USB drives detected", "dim")

    # ============================================================
    # Run All
    # ============================================================
    def _run(self):
        if self.running:
            return
        self.running = True
        self.run_btn.configure(state="disabled", text="Running...", bg="#555555")
        self._clear_log()
        self.start_time = time.time()

        thread = threading.Thread(target=self._run_thread, daemon=True)
        thread.start()
        self._update_timer()

    def _run_thread(self):
        step_map = {
            "virus": self._step_virus,
            "temp": self._step_temp,
            "recycle": self._step_recycle,
            "dns": self._step_dns,
            "hidden": self._step_hidden,
            "darkmode": self._step_darkmode,
            "telemetry": self._step_telemetry,
            "startup": self._step_startup,
            "chkdsk": self._step_chkdsk,
            "sfc": self._step_sfc,
            "dism": self._step_dism,
            "explorer": self._step_explorer,
            "usb": self._step_usb,
        }

        all_steps = STEPS + [USB_STEP]
        selected = [s for s in all_steps if self.checks[s["id"]].get()]
        total = len(selected)

        if total == 0:
            self._log("No steps selected!", "red")
            self._finish()
            return

        self._log(f"Starting {total} steps...\n", "green")
        self._log(f"Developed BY Ahmad_Khaled\n", "blue")

        for i, step in enumerate(selected):
            self.root.after(
                0,
                self._update_progress,
                i,
                total,
                f"Running: {step['name']}",
            )
            try:
                step_map[step["id"]]()
            except Exception as e:
                self._log(f"  ERROR: {e}", "red")
            self._log("", "")

        self.root.after(0, self._update_progress, total, total, "Done!")
        elapsed = time.time() - self.start_time
        self._log(f"All done in {elapsed:.1f}s", "green")
        self._log(f"BY Ahmad_Khaled", "blue")
        self._finish()

    def _finish(self):
        self.running = False
        self.root.after(
            0,
            lambda: self.run_btn.configure(
                state="normal", text="\u25B6  Run Cleaner", bg=ACCENT
            ),
        )
        self.root.after(
            0, lambda: self.step_label.configure(text="Completed", fg=ACCENT)
        )

    def _update_timer(self):
        if self.running and self.start_time:
            elapsed = time.time() - self.start_time
            m, s = divmod(int(elapsed), 60)
            self.timer_label.configure(text=f"{m:02d}:{s:02d}")
            self.root.after(1000, self._update_timer)

    def _clear_log(self):
        self.log_box.configure(state="normal")
        self.log_box.delete("1.0", "end")
        self.log_box.configure(state="disabled")
        self.log_lines.clear()

    def _export_log(self):
        if not self.log_lines:
            messagebox.showinfo("Export", "No log to export")
            return

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill

            fp = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel", "*.xlsx")],
                initialfilename=f"AK_Cleaner_Log_{datetime.datetime.now():%Y%m%d_%H%M%S}",
            )
            if not fp:
                return

            wb = Workbook()
            ws = wb.active
            ws.title = "Cleaner Log"
            ws.column_dimensions["A"].width = 10
            ws.column_dimensions["B"].width = 80
            ws.column_dimensions["C"].width = 25

            # Header row
            ws.append(["#", "Message", "Developer"])
            for col in ["A1", "B1", "C1"]:
                ws[col].font = Font(bold=True, color="FFFFFF")
                ws[col].fill = PatternFill("solid", fgColor="1A1A1A")

            color_map = {
                "green": "00C896",
                "yellow": "FFD93D",
                "red": "FF4B4B",
                "blue": "4BA3FF",
                "dim": "666666",
            }

            for i, (text, tag) in enumerate(self.log_lines, 1):
                dev = "BY Ahmad_Khaled" if i == 1 else ""
                ws.append([i, text, dev])
                if tag in color_map:
                    ws.cell(row=i + 1, column=2).font = Font(color=color_map[tag])

            # Footer credit
            last_row = len(self.log_lines) + 2
            ws.cell(row=last_row, column=2, value="Developed BY Ahmad_Khaled")
            ws.cell(row=last_row, column=2).font = Font(bold=True, color="00C896")

            wb.save(fp)
            self._log(f"Log exported to: {fp}", "green")
            messagebox.showinfo("Export", f"Log saved to:\n{fp}")

        except ImportError:
            # Fallback to .txt
            fp = filedialog.asksaveasfilename(
                defaultextension=".txt",
                filetypes=[("Text", "*.txt")],
                initialfilename=f"AK_Cleaner_Log_{datetime.datetime.now():%Y%m%d_%H%M%S}",
            )
            if fp:
                with open(fp, "w", encoding="utf-8") as f:
                    f.write("AK System Cleaner v4.3 — Developed BY Ahmad_Khaled\n")
                    f.write("=" * 50 + "\n\n")
                    for text, tag in self.log_lines:
                        f.write(text + "\n")
                    f.write("\n" + "=" * 50 + "\n")
                    f.write("Developed BY Ahmad_Khaled\n")
                self._log(f"Log exported to: {fp}", "green")
                messagebox.showinfo("Export", f"Log saved to:\n{fp}")

    # ============================================================
    # Run App
    # ============================================================
    def run(self):
        self.root.mainloop()


# ============================================================
# Main Entry Point
# ============================================================
if __name__ == "__main__":
    try:
        app = AKCleanerApp()
        app.run()
    except Exception as e:
        messagebox.showerror(
            "AK Cleaner Error",
            f"Failed to start:\n\n{traceback.format_exc()}"
        )